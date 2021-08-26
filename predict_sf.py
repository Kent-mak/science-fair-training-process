from openpyxl import load_workbook
from PIL import Image, ImageOps
import matplotlib.pyplot as plt
import tensorflow.keras
import numpy as np
import os


def predict_single (img_array,model):
    data = np.ndarray(shape=(1, 224, 224, 3), dtype=np.float32)
    normalized_array = (img_array.astype(np.float32) / 127.0) - 1
    data[0] = normalized_array
    proba = model.predict(data)
    prediction = np.argmax(proba, axis=1)
    return prediction, proba

def show_plot (image,prediction): 

    plt.gcf().set_size_inches(12,14)
    ax = plt.subplot()
    plt.imshow(image)
    title = 'prediction: Class '+str(prediction[0])
    
    ax.set_title(title,fontsize = 24)
    ax.set_xticks([])
    ax.set_yticks([])
    plt.show()

def image_to_array (image):

    size = (224, 224)
    cropped_img = ImageOps.fit(image, size, Image.ANTIALIAS)
    img_array = np.asarray(cropped_img)
    return img_array, cropped_img 
    
def to_excel (wb,ws,img_name,prediction,prob_array):
    col = ['B','C','D','E']
    ws['A{}'.format(img_name)] = img_name
    ws['F{}'.format(img_name)] = prediction 
    i = 0
    for c in col:
        ws['{}'.format(c) + '{}'.format(img_name)] \
                    = '{:.3f}'. format(prob_array[0][i]) 
        i+=1            
    wb.save('Book1.xlsx')

os.chdir(r'C:\Users\David Lee\Desktop\modelb-3\test1coded')
model = tensorflow.keras.models.load_model('keras_model2+3-1.h5')
wb = load_workbook(filename = 'Book1.xlsx')
ws = wb.worksheets[0]
for name in range(1,101):
    path = str(name) +'.jpg'    

    img = Image.open(path)
    img_array, img = image_to_array(img)

    np.set_printoptions(suppress=True)
    prediction,prob = predict_single(img_array,model)

    # predt_OneImg()

    for i in range(0,prob.shape[1]):
        print('Class{}:'.format(i)+ '{:.3f}'. format(prob[0][i]))
    print(' ')
    show_plot(img, prediction)

    to_excel(wb, ws, name, prediction[0],prob)


